import streamlit as st
import pandas as pd
import requests
import tempfile
from io import BytesIO
from openpyxl import Workbook
from openpyxl.drawing.image import Image

# Function to preprocess a single sheet
def preprocess_sheet(df):
    """Preprocess the DataFrame by removing rows above the 'TOTAL' row."""
    total_row_index = df.apply(lambda row: row.astype(str).str.contains("TOTAL", case=False, na=False)).any(axis=1).idxmax()
    processed_df = df.iloc[total_row_index:].reset_index(drop=True)

    if total_row_index != 0:
        processed_df.columns = processed_df.iloc[0].astype(str)
        processed_df = processed_df[1:].reset_index(drop=True)

    return processed_df

# Function to collect maximum marks for each sheet
def create_max_marks_UI(uploaded_file):
    xls = pd.ExcelFile(uploaded_file)
    max_marks = {}

    selected_sheets = st.multiselect("Select sheets to process:", xls.sheet_names)

    if selected_sheets:
        st.write("Set maximum marks for the selected sheets:")
        for name in selected_sheets:
            max_marks_input = st.number_input(
                f"Enter maximum marks for sheet {name}:", min_value=1, step=1, format="%d", key=f"max_marks_{name}"
            )
            max_marks[name] = max_marks_input

    return max_marks

# Function to download image from GitHub URL and save it to a temporary file
def download_image_from_github(image_url):
    response = requests.get(image_url)
    img_data = response.content

    # Save image to a temporary file
    with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp_img_file:
        tmp_img_file.write(img_data)
        return tmp_img_file.name  # Return the temporary file path

# Function to process Excel file and add image
def process_excel(uploaded_file, max_marks, metadata, image_url):
    xls = pd.ExcelFile(uploaded_file)
    result_file_path = "Result.xlsx"
    wb = Workbook()

    # Remove the default sheet that openpyxl creates when initializing a new workbook
    wb.remove(wb.active)

    # Download image from GitHub and save it to a temporary file
    image_path = download_image_from_github(image_url)
    
    # Process each sheet and add data
    for name in max_marks.keys():
        st.write(f"Processing sheet: {name}")
        df = pd.read_excel(xls, sheet_name=name)
        df = preprocess_sheet(df)
        df.columns = df.columns.str.strip().str.upper()

        if "TOTAL" not in df.columns:
            st.error(f"The column 'TOTAL' is missing in sheet '{name}'. Skipping...")
            continue

        # Get the maximum marks for this sheet
        max_mark = max_marks[name]

        # Calculate the thresholds based on 40% and 80% of the maximum marks
        weak_test_threshold = 0.40 * max_mark
        bright_test_threshold = 0.80 * max_mark

        df_weak = df[df["TOTAL"] <= weak_test_threshold]
        df_bright = df[df["TOTAL"] >= bright_test_threshold]

        df_weak.reset_index(drop=True, inplace=True)
        df_bright.reset_index(drop=True, inplace=True)

        # Adding a "Signature" column
        df_weak["SIGNATURE"] = ""
        df_bright["SIGNATURE"] = ""

        # Create a new sheet for each "Bright Students" and "Weak Students"
        ws_bright = wb.create_sheet(f"Bright Students - {name}")
        ws_weak = wb.create_sheet(f"Weak Students - {name}")

        # Insert the image at the top of each sheet
        img_bright = Image(image_path)
        ws_bright.add_image(img_bright, 'A1')
        
        img_weak = Image(image_path)
        ws_weak.add_image(img_weak, 'A1')

        # Add metadata below the image for "Bright Students" sheet
        start_row_bright = img_bright.height // 15 + 5  # Adjust based on the image height
        start_row_weak = img_weak.height // 15 + 5  # Adjust based on the image height
        
        # Create metadata info
        metadata_info = pd.DataFrame(
            {
                "Metadata": ["Name of the Faculty:", "Program:", "Class:", "Year:", "Semester:", "Course:", "Course Code:"],
                "Details": [metadata["faculty"], metadata["program"], metadata["class"], metadata["year"], metadata["semester"], metadata["course"], metadata["course_code"]],
            }
        )

        # Write metadata to "Bright Students" sheet
        for i, row in metadata_info.iterrows():
            ws_bright[f"A{start_row_bright + i}"] = row['Metadata']
            ws_bright[f"B{start_row_bright + i}"] = row['Details']

        # Write metadata to "Weak Students" sheet
        for i, row in metadata_info.iterrows():
            ws_weak[f"A{start_row_weak + i}"] = row['Metadata']
            ws_weak[f"B{start_row_weak + i}"] = row['Details']

        # Insert a line between metadata and the data (Weak/Bright Students)
        blank_row_bright = start_row_bright + len(metadata_info) + 1
        blank_row_weak = start_row_weak + len(metadata_info) + 1

        ws_bright[f"A{blank_row_bright}"] = ""  # Blank row

        ws_weak[f"A{blank_row_weak}"] = ""  # Blank row

        # Add "List of Weak Students" or "List of Bright Students" line
        weak_bright_line = "List of Weak Students" if "Weak" in name else "List of Bright Students"
        ws_bright[f"A{blank_row_bright + 1}"] = weak_bright_line
        ws_weak[f"A{blank_row_weak + 1}"] = weak_bright_line

        # Write the data to the "Bright Students" sheet
        for i, col in enumerate(df_bright.columns, 1):
            ws_bright.cell(row=blank_row_bright + 2, column=i, value=col)

        for i, row in enumerate(df_bright.itertuples(), blank_row_bright + 3):
            for j, value in enumerate(row[1:], 1):  # Skip the index
                ws_bright.cell(row=i, column=j, value=value)

        # Write the data to the "Weak Students" sheet
        for i, col in enumerate(df_weak.columns, 1):
            ws_weak.cell(row=blank_row_weak + 2, column=i, value=col)

        for i, row in enumerate(df_weak.itertuples(), blank_row_weak + 3):
            for j, value in enumerate(row[1:], 1):  # Skip the index
                ws_weak.cell(row=i, column=j, value=value)

    # Save the workbook
    wb.save(result_file_path)

    return result_file_path

# Streamlit App
st.title("Test Scoring App")

uploaded_file = st.file_uploader("Upload an Excel file", type=["xlsx"])

if uploaded_file:
    st.write("Excel file uploaded successfully!")

    # Include the new "Program" dropdown
    metadata = {
        "faculty": st.text_input("Name of Faculty"),
        "program": st.selectbox(
            "Program", 
            [
                "Electronics and Telecommunication Engg", 
                "Information Technology", 
                "Computer Engineering", 
                "Mechanical Engineering", 
                "Computer Science and Engineering (Data Science)",
                "Artificial Intelligence and Machine Learning", 
                "Artificial Intelligence (AI) and Data Science", 
                "Computer Science and Engineering (IOT and Cyber Security with Block Chain Technology)"
            ], 
            key="program_selector"
        ),  # Program dropdown
        "class": st.selectbox("Select Class", ["F.Y. B.TECH", "S.Y. B.TECH", "T.Y. B.TECH", "Final Year B.TECH"], key="class_selector"),
        "year": st.text_input("Year"),  # Year field (independent)
        "semester": st.selectbox(
            "Select Semester",
            {
                "F.Y. B.TECH": ["I", "II"],
                "S.Y. B.TECH": ["III", "IV"],
                "T.Y. B.TECH": ["V", "VI"],
                "Final Year B.TECH": ["VII", "VIII"]
            }[st.session_state.get("class_selector", "F.Y. B.TECH")]  # Dynamically fetch the semester options based on selected class
        ),
        "course": st.text_input("Course"),  # Course field
        "course_code": st.text_input("Course Code"),  # Course Code field
    }


    # Set the GitHub URL for the image (public raw image URL)
    image_url = "https://raw.githubusercontent.com/Agam-Saraf/TestScorerApp/551c9d2ad4930bafb67d8a5af567a1c7e46c9ff7/img.png"  # Replace with your actual image URL

    max_marks = create_max_marks_UI(uploaded_file)

    if max_marks and st.button("Start Processing"):
        if all(val is not None and val > 0 for val in max_marks.values()):
            result_path = process_excel(uploaded_file, max_marks, metadata, image_url)

            # Excel Download
            with open(result_path, "rb") as f:
                st.download_button(
                    label="Download Processed Excel File",
                    data=f,
                    file_name="Result.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
        else:
            st.error("Please enter valid maximum marks for all sheets.")